from decimal import Decimal, ROUND_DOWN
import re
from bs4 import BeautifulSoup
import pyautogui
import time
import tkinter as tk
from tkinter import simpledialog
import os
import pandas as pd
import openpyxl

# Função para obter o código da empresa, o mês e o ano do usuário
def get_user_input():
    root = tk.Tk()
    root.withdraw()  # Esconder a janela principal

    # Solicitar o código da empresa
    company_code = simpledialog.askstring(title="Código da Empresa", prompt="Digite o código da empresa:")

    # Solicitar o mês e o ano
    month_year = simpledialog.askstring(title="Mês e Ano", prompt="Digite o mês e o ano (MMYYYY):")

    return company_code, month_year

# Obter o código da empresa e o mês e o ano do usuário
company_code, month_year = get_user_input()

# Criar a variável day_month_year
day_month_year = '01' + month_year

# Caminho para o executável na rede
network_path = r'\\ap05\modulos\UNICO.EXE'

# Caminho para salvar os CSVs
# output_csv_path = r'C:\FOLHA\data\output\planilhas\relatorio_ferias.csv'
output_html_path_13 = r'C:\FOLHA\data\output\planilhas\relatorio_13'

# Pressionar Win + R
pyautogui.hotkey('win', 'r')
time.sleep(1)  # Esperar um momento para a janela de execução abrir

# Digitar o caminho do executável na rede
pyautogui.typewrite(network_path)
pyautogui.press('enter')

# Esperar 12 segundos para o executável carregar
time.sleep(12)

# Digitar "contabil"
pyautogui.typewrite('contabil')

# Pressionar "tab"
pyautogui.press('tab')

# Digitar "1234"
pyautogui.typewrite('1234')

# Pressionar "enter"
pyautogui.press('enter')

# Esperar um momento para a próxima ação
time.sleep(5)

# Pressionar "Ctrl + 0"
pyautogui.hotkey('ctrl', '0')

# Esperar 10 segundos para a próxima ação
time.sleep(10)

# Pressionar "Alt + R", "O", "F"
pyautogui.hotkey('alt', 'r')
time.sleep(0.5)
pyautogui.press('o')
time.sleep(0.5)
pyautogui.press('f')

# Esperar um momento para a próxima ação
time.sleep(1)

# Digitar o código da empresa
pyautogui.typewrite(company_code)
pyautogui.press('enter')
time.sleep(1)
# Pressionar "tab"
pyautogui.press('tab')

# Digitar o mês e o ano no formato MMYYYY
pyautogui.typewrite(month_year)
pyautogui.press('enter')

# Mover o clique
pyautogui.click(x=100, y=165)
time.sleep(1)

pyautogui.click(x=100, y=228)
time.sleep(5)

# Pressionar "enter" para salvar o CSV
pyautogui.press('enter')
time.sleep(5)

# Caminho do arquivo HTML
html_file_path = r"C:\FOLHA\data\output\planilhas\relatorio_ferias.htm"

# Digitar o caminho para salvar o CSV
pyautogui.typewrite(html_file_path)

pyautogui.press('enter')
time.sleep(5)
pyautogui.hotkey('alt', 'tab')

# Aguardar um momento para garantir que o arquivo seja salvo
time.sleep(5)

# Abrir e ler o conteúdo do arquivo HTML
with open(html_file_path, 'r', encoding='utf-8') as file:
    html_content = file.read()

# Usar BeautifulSoup para fazer o parsing do HTML
soup = BeautifulSoup(html_content, 'html.parser')

# Procurar pela tag <td> com o conteúdo específico "(=) Saldo final da conta"
td_elemento = soup.find('td', string="(=) Saldo final da conta")

# Verificar se o conteúdo foi encontrado
if td_elemento:
    # Encontrar todos os <td> na mesma linha após o <td> encontrado
    td_siguentes = td_elemento.find_parent('tr').find_all('td')
    
    # Função para converter os valores para float e usar ',' como separador decimal
    def converter_valor(valor):
        valor = valor.replace('.', '').replace(',', '.')
        return float(valor)
    
    # Extrair os valores da linha, ignorando o segundo dado 'R$'
    saldo_final_conta = td_siguentes[0].get_text(strip=True)
    ferias = converter_valor(td_siguentes[2].get_text(strip=True))
    fgts_f = converter_valor(td_siguentes[3].get_text(strip=True))
    inss_f = converter_valor(td_siguentes[4].get_text(strip=True))
    pis_f = converter_valor(td_siguentes[5].get_text(strip=True))
    total_f = converter_valor(td_siguentes[6].get_text(strip=True))
    
    # Exibir os valores extraídos
    print(f"Saldo final da conta: {saldo_final_conta}")
    print(f"Férias: {ferias}")
    print(f"FGTS_f: {fgts_f}")
    print(f"INSS_f: {inss_f}")
    print(f"PIS_f: {pis_f}")
    print(f"Total_f: {total_f}")
else:
    print("Elemento específico não encontrado.")

# Verificar se o arquivo foi salvo corretamente
# if os.path.exists(output_csv_path):
#     print(f"Arquivo CSV salvo em: {output_csv_path}")
# else:
#     print("Erro ao salvar o arquivo CSV.")

# Ler o CSV e procurar por "(=) Saldo final da conta" na coluna A
# df = pd.read_csv(output_csv_path, sep=';', encoding='latin1', on_bad_lines='warn')

# Substituir os nomes das colunas conforme solicitado
# df.rename(columns={
#     'Unnamed: 5': 'FERIAS',
#     'Unnamed: 11': 'FGTS_f',
#     'Unnamed: 16': 'INSS_f',
#     'Unnamed: 23': 'PIS_f',
#     'Unnamed: 29': 'TOTAL_f'
# }, inplace=True)

# Procurar por "(=) Saldo final da conta" na coluna A
# saldo_final_row = df[df.iloc[:, 0].str.contains(r"\(=\) Saldo final da conta", na=False)]

# Variáveis para armazenar os dados
saldo_final_valores = {
    'FERIAS': ferias,
    'FGTS_f': fgts_f,
    'INSS_f': inss_f,
    'PIS_f': pis_f,
    'TOTAL_f': total_f
}

# Se encontrado, armazenar os valores correspondentes à linha
# if not saldo_final_row.empty:
#     saldo_final_valores = saldo_final_row[['FERIAS', 'FGTS_f', 'INSS_f', 'PIS_f', 'TOTAL_f']].iloc[0].dropna().to_dict()
#     print("Saldo final da conta encontrado. Valores correspondentes:")
#     print(saldo_final_valores)
# else:
#     print("Saldo final da conta não encontrado.")

# Converter os valores de saldo_final_valores para float
# if saldo_final_valores:
#     for key in saldo_final_valores:
#         saldo_final_valores[key] = float(saldo_final_valores[key].replace('.', '').replace(',', '.'))

# Converter o valor de INSS_f para número
# if saldo_final_valores and 'INSS_f' in saldo_final_valores:
#     inss_f_valor = saldo_final_valores['INSS_f']
#     print(f"Valor de INSS_f: {inss_f_valor}")

# Salvar os valores encontrados em um novo CSV
# if saldo_final_valores:
#     output_csv_path_final = r'C:\FOLHA\data\output\planilhas\saldo_final_valores.csv'
#     pd.DataFrame([saldo_final_valores]).to_csv(output_csv_path_final, index=False, sep=';')
#     print(f"Arquivo CSV com valores encontrados salvo em: {output_csv_path_final}")

# Repetir o processo para relatorio_13.csv
# Pressionar "Alt + R", "O", "1"
pyautogui.hotkey('alt', 'r')
time.sleep(0.5)
pyautogui.press('o')
time.sleep(0.5)
pyautogui.press('1')

time.sleep(2)

# Pressionar "tab" e digitar o mês
pyautogui.press('tab')
pyautogui.typewrite(month_year)
pyautogui.press('tab')

# Mover o clique
pyautogui.click(x=100, y=165)
time.sleep(1)

pyautogui.click(x=100, y=228)
time.sleep(5)

# Pressionar "enter" para salvar o html 
pyautogui.press('enter')    
time.sleep(2)
pyautogui.write(output_html_path_13)
pyautogui.press('enter')

# Aguardar um momento para garantir que o arquivo seja salvo
time.sleep(5)

# Caminho do seu arquivo HTML
caminho_arquivo = output_html_path_13

# Função para formatar os números no formato desejado
def formatar_valor(valor):
    # Substituindo a vírgula por ponto (para garantir que a conversão para float funcione)
    valor = valor.replace('.', '').replace(',', '.')
    try:
        # Convertendo para float e depois formatando para a string no formato desejado
        return '{:,.2f}'.format(float(valor)).replace(',', 'X').replace('.', ',').replace('X', '.')
    except ValueError:
        return valor  # Retorna o valor original se não for um número válido

# Abrindo o arquivo HTML
with open(r"C:\FOLHA\data\output\planilhas\relatorio_13.htm", 'r', encoding='utf-8') as file:
    html_content = file.read()

# Usando BeautifulSoup para fazer o parsing do HTML
soup = BeautifulSoup(html_content, 'html.parser')

# Procurando pela tag <td> com o conteúdo específico
td_elemento = soup.find('td', string="(=) Saldo final da conta")

# Verificando se o conteúdo foi encontrado
if td_elemento:
    # Encontrando todos os <td> na mesma linha após o <td> encontrado
    td_siguentes = td_elemento.find_all_next('td')
    
    # Definindo os rótulos personalizados para os 6 primeiros valores
    rotulos = [
        "salario_13", 
        "fgts_13", 
        "inss_13", 
        "pis_13", 
        "Total_13"
    ]
    
    # Exibindo os 6 primeiros valores com rótulos personalizados
    valores_13 = {}
    for i, td in enumerate(td_siguentes[1:6]):
        valores_13[rotulos[i]] = td.get_text()
        print(f'{rotulos[i]}: {td.get_text()}')
else:
    print("Elemento específico não encontrado.")

# Transformar o valor de salario_13 no estilo 88340.34
if 'salario_13' in valores_13:
    salario_13_formatado = valores_13['salario_13'].replace('.', '').replace(',', '.')
    valores_13['salario_13'] = salario_13_formatado
    print(f"salario_13 formatado: {salario_13_formatado}")

# Transformar o valor de inss_13 no estilo 88340.34
if 'inss_13' in valores_13:
    inss_13_formatado = valores_13['inss_13'].replace('.', '').replace(',', '.')
    valores_13['inss_13'] = inss_13_formatado
    print(f"inss_13 formatado: {inss_13_formatado}")

# Transformar o valor de fgts_13 no estilo 88340.34
if 'fgts_13' in valores_13:
    fgts_13_formatado = valores_13['fgts_13'].replace('.', '').replace(',', '.')
    valores_13['fgts_13'] = fgts_13_formatado
    print(f"fgts_13 formatado: {fgts_13_formatado}")

time.sleep(5)

pyautogui.hotkey('alt','tab')
time.sleep(0.5)
pyautogui.press('alt')
time.sleep(0.5)
pyautogui.press('r')
time.sleep(0.5)
pyautogui.press('i')
time.sleep(0.5)
pyautogui.press('b')
time.sleep(0.5)
pyautogui.press('i')
# Esperar um momento para a próxima ação
time.sleep(4)

time.sleep(1)
pyautogui.press('tab')
time.sleep(1)
pyautogui.press('tab')
time.sleep(1)
pyautogui.press('tab')
time.sleep(1)
pyautogui.press('tab')
time.sleep(1)
# Digitar o mês e o ano no formato MMYYYY
pyautogui.typewrite(month_year)
time.sleep(1)
pyautogui.press('enter')
time.sleep(1)
pyautogui.press('tab')

# Mover o clique
pyautogui.click(x=100, y=165)
time.sleep(1)

pyautogui.click(x=100, y=228)
time.sleep(5)

output_html_path_inss = r'C:\FOLHA\data\output\planilhas\relatorio_inss'

# Pressionar "enter" para salvar o html
pyautogui.press('enter')
time.sleep(5)
pyautogui.write(output_html_path_inss)
pyautogui.press('enter')

time.sleep(5)
with open(r"C:\FOLHA\data\output\planilhas\relatorio_inss.htm", 'r', encoding='utf-8') as file:
    html_content_inss = file.read()
soup = BeautifulSoup(html_content_inss, 'html.parser')

def extrair_valor(valor):
    # Remover qualquer texto extra após o número
    valor = re.sub(r'\s\(.*\)', '', valor)  # Remove qualquer texto dentro de parênteses e o próprio parêntese
    # Remover pontos de milhar e substituir vírgula por ponto para a conversão de número
    valor = valor.replace('.', '').replace(',', '.')
    return float(valor)

# Encontrar todas as células <td> com a classe 's40'
td_s40_values = soup.find_all('td', class_='s40')

# Extrair e formatar os valores
valores = []
for td in td_s40_values:
    valor = td.get_text(strip=True)  # Extrai o texto da célula
    valor_formatado = extrair_valor(valor)  # Formatar o valor extraído
    valores.append(valor_formatado)

variaveis = {}

# Exibir os valores formatados
for i, valor in enumerate(valores):
    # Aqui estamos associando a chave "INSS" ao valor
    variaveis["INSS"] = valor  # Armazena o valor em variaveis['INSS']

# Exibindo o valor de INSS
print(f"INSS {variaveis['INSS']}")

# Imprimir o caminho do ambiente virtual
print("Caminho do ambiente virtual: c:/FOLHA/.venv/Scripts/python.exe")

time.sleep(4)
pyautogui.hotkey('alt','tab')
time.sleep(0.5)
pyautogui.press('alt')
time.sleep(0.5)
pyautogui.press('r')
time.sleep(0.5)
pyautogui.press('i')
time.sleep(0.5)
pyautogui.press('b')
time.sleep(0.5)
pyautogui.press('f')

time.sleep(2)

pyautogui.press('tab')
pyautogui.typewrite(month_year)
# Esperar um momento para a próxima ação

# Mover o clique
pyautogui.click(x=100, y=165)
time.sleep(1)

pyautogui.click(x=100, y=228)
time.sleep(5)

output_html_path_fgts = r'C:\FOLHA\data\output\planilhas\relatorio_fgts'

# Caminho do arquivo HTML
caminho_html_sal_pro = r"C:\FOLHA\data\output\planilhas\relatorio_sal_pro.htm"

# Pressionar "enter" para salvar o html
pyautogui.press('enter')
time.sleep(5)
pyautogui.write(output_html_path_fgts)
pyautogui.press('enter')
time.sleep(3)

pyautogui.hotkey('alt','tab')
time.sleep(1)
pyautogui.press('alt')
time.sleep(1)
pyautogui.press('r')
time.sleep(1)
pyautogui.press('e')
time.sleep(5)
pyautogui.press('tab')
time.sleep(1)
pyautogui.press('tab')
time.sleep(1)
pyautogui.press('tab')
time.sleep(1)
pyautogui.write(month_year)
time.sleep(2)
pyautogui.leftClick(94, 165)
time.sleep(1)
pyautogui.leftClick(95, 228)
time.sleep(3)
pyautogui.press('enter')
time.sleep(2)
pyautogui.write(caminho_html_sal_pro)
pyautogui.press('enter')
time.sleep(3)

with open(output_html_path_fgts +'.htm') as file:
    html_content_inss = file.read()
soup = BeautifulSoup(html_content_inss, 'html.parser')

def extrair_valor(valor_fgts):
    # Remover qualquer texto extra após o número
    valor = re.sub(r'\s\(.*\)', '', valor)  # Remove qualquer texto dentro de parênteses e o próprio parêntese
    # Remover pontos de milhar e substituir vírgula por ponto para a conversão de número
    valor = valor.replace('.', '').replace(',', '.')
    return float(valor)


td_s12_values = soup.find_all('td', {'colspan': '3', 'class': 's12'})
if td_s12_values:
    # Obter o último valor
    valor_bruto = td_s12_values[-1].text  # Último valor

    # Formatar o valor para xxxxx,xx
    valor_formatado_fgts = valor_bruto.replace('.', '').replace(',', '.')  # Remove os pontos e ajusta vírgula
    print("Valor formatado:", valor_formatado_fgts)

else:
    print("Nenhum valor encontrado.")


# Caminho da planilha de conciliação
conciliacao_path = f'C:\\projeto\\planilhas\\balancete\\CONCILIACAO_{company_code}_{month_year}.xlsx'

# Verificar se o arquivo de conciliação existe e é válido
if not os.path.exists(conciliacao_path):
    print(f"Erro: O arquivo {conciliacao_path} não existe.")
else:
    try:
        # Abrir a planilha de conciliação
        wb = openpyxl.load_workbook(conciliacao_path)

        # Atualizar a Sheet1
        ws1 = wb.active

        # Mapeamento dos valores da coluna H com os valores da coluna A
        valores_mapeados = {}
        for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=8):
            cell_a = row[0]  # Coluna A
            cell_h = row[7]  # Coluna H
            if cell_a.value is not None:
                valores_mapeados[cell_a.value] = cell_h.value

        # Procurar os valores e retornar os valores da coluna H
        vlr_prov_ferias = valores_mapeados.get(177, None)
        vlr_prov_fgts_f = valores_mapeados.get(179, None)
        vlr_prov_inss_f = valores_mapeados.get(181, None)
        vlr_prov_13 = valores_mapeados.get(176, None)
        vlr_prov_inss_13 = valores_mapeados.get(180, None)
        vlr_prov_fgts_13 = valores_mapeados.get(178, None)
        vlr_INSS = valores_mapeados.get(171, None)
        vlr_fgts = valores_mapeados.get(172, None)

        if vlr_fgts is not None and vlr_fgts != "Valor não encontrado":
            vlr_fgts = float(vlr_fgts)
            valor_formatado_fgts = float(valor_formatado_fgts)
            vlr_fgts = Decimal(vlr_fgts)
            vlr_fgts = vlr_fgts.quantize(Decimal('0.00'), rounding=ROUND_DOWN)

            if vlr_fgts == valor_formatado_fgts:
                print("Os valores são iguais.")
            else:
                print("Os valores são diferentes.")
        else:
            print("vlr_fgts não encontrado na planilha.")

        print(f"vlr_prov_ferias: {vlr_prov_ferias}")
        print(f"vlr_prov_fgts_f: {vlr_prov_fgts_f}")
        print(f"vlr_prov_inss_f: {vlr_prov_inss_f}")
        print(f"vlr_prov_13: {vlr_prov_13}")
        print(f"vlr_prov_inss_13: {vlr_prov_inss_13}")
        print(f"vlr_prov_fgts_13: {vlr_prov_fgts_13}")
        print(f"vlr_INSS: {vlr_INSS}")
        print(f"vlr_fgts: {vlr_fgts}")

        # Comparar os valores
        try:
            salario_13_float = float(salario_13_formatado)
            vlr_prov_13_float = float(vlr_prov_13) if vlr_prov_13 is not None and vlr_prov_13 != "Valor não encontrado" else None
            resultado_13 = "OK" if salario_13_float == vlr_prov_13_float else "verificar"
        except (ValueError, TypeError) as e:
            print(f"Erro ao converter valores para float: {e}")
            resultado_13 = "verificar"

        try:
            inss_13_float = float(inss_13_formatado)
            vlr_prov_inss_13_float = float(vlr_prov_inss_13) if vlr_prov_inss_13 is not None and vlr_prov_inss_13 != "Valor não encontrado" else None
            resultado_inss_13 = "OK" if inss_13_float == vlr_prov_inss_13_float else "verificar"
        except (ValueError, TypeError) as e:
            print(f"Erro ao converter valores para float: {e}")
            resultado_inss_13 = "verificar"

        try:
            fgts_13_float = float(fgts_13_formatado)
            vlr_prov_fgts_13_float = float(vlr_prov_fgts_13) if vlr_prov_fgts_13 is not None and vlr_prov_fgts_13 != "Valor não encontrado" else None
            resultado_fgts_13 = "OK" if fgts_13_float == vlr_prov_fgts_13_float else "verificar"
        except (ValueError, TypeError) as e:
            print(f"Erro ao converter valores para float: {e}")
            resultado_fgts_13 = "verificar"

        if vlr_INSS is not None and vlr_INSS != "Valor não encontrado":
            try:
                inss_float = float(variaveis["INSS"])
                vlr_prov_inss_float = float(vlr_INSS)
                resultado_inss = "OK" if inss_float == vlr_prov_inss_float else "verificar"
            except (ValueError, TypeError) as e:
                print(f"Erro ao converter valores para float: {e}")
                resultado_inss = "verificar"
        else:
            resultado_inss = "verificar"
            print("vlr_INSS não encontrado na planilha.")

        resultado_ferias = "verificar"
        resultado_fgts_f = "verificar"
        resultado_inss_f = "verificar"

        if saldo_final_valores:
            resultado_ferias = "OK" if saldo_final_valores.get('FERIAS') == vlr_prov_ferias else "verificar"
            resultado_fgts_f = "OK" if saldo_final_valores.get('FGTS_f') == vlr_prov_fgts_f else "verificar"
            resultado_inss_f = "OK" if saldo_final_valores.get('INSS_f') == vlr_prov_inss_f else "verificar"

        resultado_fgts = "OK" if float(valor_formatado_fgts) == float(vlr_fgts) else "verificar"

        print(f"Comparação FERIAS: {resultado_ferias}")
        print(f"Comparação FGTS_f: {resultado_fgts_f}")
        print(f"Comparação INSS_f: {resultado_inss_f}")
        print(f"Comparação salario_13: {resultado_13}")
        print(f"Comparação inss_13: {resultado_inss_13}")
        print(f"Comparação fgts_13: {resultado_fgts_13}")
        print(f"Comparação INSS: {resultado_inss}")
        print(f"Comparação FGTS: {resultado_fgts}")

        # Escrever os resultados na coluna I ao lado dos valores
        for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == 177:
                    ws1.cell(row=cell.row, column=9, value=resultado_ferias)
                elif cell.value == 179:
                    ws1.cell(row=cell.row, column=9, value=resultado_fgts_f)
                elif cell.value == 181:
                    ws1.cell(row=cell.row, column=9, value=resultado_inss_f)
                elif cell.value == 176:
                    ws1.cell(row=cell.row, column=9, value=resultado_13)
                elif cell.value == 180:
                    ws1.cell(row=cell.row, column=9, value=resultado_inss_13)
                elif cell.value == 178:
                    ws1.cell(row=cell.row, column=9, value=resultado_fgts_13)
                elif cell.value == 171:
                    ws1.cell(row=cell.row, column=9, value=resultado_inss)
                elif cell.value == 172:
                    ws1.cell(row=cell.row, column=9, value=resultado_fgts)

        # Salvar a planilha novamente
        wb.save(conciliacao_path)

        print(f"Resultados escritos na planilha de conciliação.")
    except Exception as e:
        print(f"Erro ao abrir ou processar o arquivo de conciliação: {e}")

# Função para ler o HTML de um caminho especificado
def ler_html(caminho_html):
    with open(caminho_html, 'r', encoding='utf-8') as file:
        html_content = file.read()
    return html_content

# Função para procurar a linha onde o primeiro valor é "Líquido"
def procurar_linha_liquido(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    linha_liquido = None

    for td in soup.find_all('td'):
        if td.get_text(strip=True) == "Líquido":
            linha_liquido = td.find_parent('tr')
            break

    return linha_liquido

# Ler o conteúdo do HTML
html_content_sal_pro = ler_html(caminho_html_sal_pro)

# Procurar e salvar a linha onde o primeiro valor é "Líquido"
linha_liquido = procurar_linha_liquido(html_content_sal_pro)
if linha_liquido:
    colunas = linha_liquido.find_all('td')
    resultado_geral = colunas[1].get_text(strip=True).replace('.', '').replace(',', '.')
    colaboradores = colunas[3].get_text(strip=True).replace('.', '').replace(',', '.')
    empregadores = colunas[5].get_text(strip=True).replace('.', '').replace(',', '.')
    autonomos = colunas[7].get_text(strip=True).replace('.', '').replace(',', '.')
    estagiarios = colunas[9].get_text(strip=True).replace('.', '').replace(',', '.')

    print("Resultado Geral:", resultado_geral)
    print("Colaboradores:", colaboradores)
    print("Empregadores:", empregadores)
    print("Autônomos:", autonomos)
    print("Estagiários:", estagiarios)
else:
    print("Linha com 'Líquido' não encontrada.")

# Caminho da planilha de conciliação
conciliacao_path = f'C:\\projeto\\planilhas\\balancete\\CONCILIACAO_{company_code}_{month_year}.xlsx'

# Verificar se o arquivo de conciliação existe e é válido
if not os.path.exists(conciliacao_path):
    print(f"Erro: O arquivo {conciliacao_path} não existe.")
else:
    try:
        # Abrir a planilha de conciliação
        wb = openpyxl.load_workbook(conciliacao_path)
        ws = wb.active

        # Procurar o valor na coluna A e comparar com os valores de colaboradores e empregadores
        for row in ws.iter_rows(min_row=2):
            cell_a = row[0].value  # Coluna A (índice 0)
            cell_h = row[7].value  # Coluna H (índice 7)
            cell_i = row[8]        # Coluna I (índice 8)

            if cell_a == 160:
                if float(cell_h) == float(colaboradores):
                    cell_i.value = "OK"
                else:
                    cell_i.value = "Verificar"
            elif cell_a == 169:
                if float(cell_h) == float(empregadores):
                    cell_i.value = "OK"
                else:
                    cell_i.value = "Verificar"

        # Salvar as alterações de volta no arquivo Excel
        wb.save(conciliacao_path)
        print(f"Resultado da comparação escrito na planilha de conciliação.")
    except Exception as e:
        print(f"Erro ao abrir ou processar o arquivo de conciliação: {e}")


